# -*- coding: utf-8 -*-
# created By: Williams Sousa
# created Date: 20/02/2023
# version: '1.6.0'
# Type: Python3.9+

"""" script para crawler em site """

'----------imports-------------'
import requests
import json

from openpyxl import Workbook ,load_workbook
from threading import Thread , active_count
from time import sleep

'-------------CODE!---------------'

url='https://api.osf.io/v2/preprints/{}/?metrics%5Bdownloads%5D=total&metrics%5Bviews%5D=total'
n_thread= active_count()

def get_in_url(i):
	global master
	link=i.split('/')[3]
	resp=requests.get(url.format(link))
	if resp.status_code == 200:
		meta=json.loads(resp.text)['meta']['metrics']
		master[i]['metrics']=meta

def salv():
	from_write=[[i,
			master[i]['metrics']['downloads'],
			master[i]['metrics']['views']]  for i in master]
	from_write.insert(0,['links','downloads','views'])

	pp=Workbook()
	page=pp.active
	for i in from_write:
		page.append(i)
	pp.save(link_file)
	return 1


def load_var():
	global max_thread, sleep_time, time_max,time_max,master,link_file
	with open('app.conf','r',encoding='utf-8') as arq:
		conf=json.loads(arq.read())
	max_thread=conf['conf']['threads'] + n_thread
	sleep_time=conf['conf']['sleep']
	time_max=conf['conf']['time_max']
	link_file=conf['conf']['link_file']

	p=load_workbook(link_file).active
	dados=[i.value for i in p['a']]
	dados.pop(0)
	master = {i:{'metrics':{}} for i in dados}


def main():
	global total
	n=0
	q=len(master)
	for i in master:
		if active_count() > max_thread:
			sleep(sleep_time)
		total=q-n
		Thread(target=get_in_url,args=(i,)).start()
		n+=1


if __name__=='__main__':
	print(f'coletando...'.upper())
	load_var()
	Thread(target=main).start()
	n=0
	while (active_count() > n_thread) & (n < time_max):
		print(f'processos..{active_count()-n_thread:.>5}      restantes{total:.>5}      {100-total/len(master)*100:.1f}%')
		sleep(0.5)
		n+=0.5
	print('salvando...')
	if salv():
		input('salvo com sucesso.'.upper())
	else:
		print('erro em salvar!')
	sleep(1)